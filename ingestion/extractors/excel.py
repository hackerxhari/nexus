"""
Excel extractor — handles .xlsx and .xls files.
Key feature: Row-column relationship extraction.
First row treated as headers, data rows stored as Header:Value pairs
so the embedding model captures meaningful semantic relationships.

Uses openpyxl (.xlsx) and xlrd (.xls) — both already in requirements.
"""

import os
from typing import List, Optional

from core.exceptions import ExtractionFailedError
from core.logger import get_logger
from ingestion.extractors.base import BaseExtractor, ExtractionResult

logger = get_logger(__name__)


class ExcelExtractor(BaseExtractor):

    @property
    def supported_extensions(self) -> List[str]:
        return ["xlsx", "xls"]

    def _extract(self, filepath: str) -> ExtractionResult:
        if not os.path.exists(filepath):
            raise ExtractionFailedError(
                os.path.basename(filepath),
                "File not found"
            )

        ext = filepath.rsplit(".", 1)[-1].lower()

        try:
            if ext == "xls":
                return self._extract_xls(filepath)
            else:
                return self._extract_xlsx(filepath)
        except ExtractionFailedError:
            raise
        except Exception as e:
            raise ExtractionFailedError(
                os.path.basename(filepath),
                f"Excel parsing failed: {str(e)}"
            )

    def _extract_xlsx(self, filepath: str) -> ExtractionResult:
        """Extract from .xlsx files using openpyxl."""
        from openpyxl import load_workbook

        warnings = []
        sheet_sections = []
        total_rows = 0

        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            sheet_count = len(wb.sheetnames)

            if sheet_count == 0:
                raise ExtractionFailedError(
                    os.path.basename(filepath),
                    "Workbook has no sheets"
                )

            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]
                    sheet_text = self._extract_sheet(ws, sheet_name)
                    if sheet_text:
                        sheet_sections.append(sheet_text["text"])
                        total_rows += sheet_text["row_count"]
                except Exception as e:
                    warnings.append(
                        f"Sheet '{sheet_name}' error: {str(e)}"
                    )
                    continue

            wb.close()

        except ExtractionFailedError:
            raise
        except Exception as e:
            raise ExtractionFailedError(
                os.path.basename(filepath),
                f"XLSX parsing failed: {str(e)}"
            )

        combined_text = "\n\n".join(sheet_sections)

        if not combined_text.strip():
            warnings.append("No data found in workbook")

        return ExtractionResult(
            text=combined_text,
            page_count=sheet_count,
            extraction_method="xlsx_openpyxl",
            warnings=warnings,
            metadata={
                "sheet_count": sheet_count,
                "total_data_rows": total_rows
            }
        )

    def _extract_xls(self, filepath: str) -> ExtractionResult:
        """Extract from legacy .xls files using xlrd."""
        import xlrd

        warnings = []
        sheet_sections = []
        total_rows = 0

        try:
            wb = xlrd.open_workbook(filepath)
            sheet_count = wb.nsheets

            if sheet_count == 0:
                raise ExtractionFailedError(
                    os.path.basename(filepath),
                    "Workbook has no sheets"
                )

            for sheet_idx in range(sheet_count):
                try:
                    ws = wb.sheet_by_index(sheet_idx)
                    sheet_text = self._extract_xls_sheet(ws)
                    if sheet_text:
                        sheet_sections.append(sheet_text["text"])
                        total_rows += sheet_text["row_count"]
                except Exception as e:
                    warnings.append(
                        f"Sheet {sheet_idx} error: {str(e)}"
                    )
                    continue

        except ExtractionFailedError:
            raise
        except Exception as e:
            raise ExtractionFailedError(
                os.path.basename(filepath),
                f"XLS parsing failed: {str(e)}"
            )

        combined_text = "\n\n".join(sheet_sections)

        if not combined_text.strip():
            warnings.append("No data found in workbook")

        return ExtractionResult(
            text=combined_text,
            page_count=sheet_count,
            extraction_method="xls_xlrd",
            warnings=warnings,
            metadata={
                "sheet_count": sheet_count,
                "total_data_rows": total_rows
            }
        )

    def _extract_sheet(self, ws, sheet_name: str) -> Optional[dict]:
        """
        Extract a sheet from an openpyxl worksheet.
        Uses first row as headers for row-column relationship structuring.

        Example output:
            [Sheet: Employee Data]
            Row 1: Name: John | Department: HR | Salary: 50000
            Row 2: Name: Jane | Department: IT | Salary: 60000
        """
        rows_data = []
        headers = []
        row_count = 0

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            # Skip completely empty rows
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            cells = [
                str(cell).strip() if cell is not None else ""
                for cell in row
            ]

            if row_idx == 0:
                # First non-empty row as headers
                headers = cells
                continue

            # Format data rows as Header: Value pairs
            row_count += 1
            formatted_row = self._format_data_row(
                headers, cells, row_count
            )
            if formatted_row:
                rows_data.append(formatted_row)

        if not rows_data:
            return None

        header_line = f"[Sheet: {sheet_name}]"
        if headers and any(h for h in headers):
            header_line += f"\nColumns: {' | '.join(h for h in headers if h)}"

        return {
            "text": f"{header_line}\n" + "\n".join(rows_data),
            "row_count": row_count
        }

    def _extract_xls_sheet(self, ws) -> Optional[dict]:
        """Extract a sheet from an xlrd worksheet."""
        rows_data = []
        headers = []
        row_count = 0

        for row_idx in range(ws.nrows):
            cells = [
                str(ws.cell_value(row_idx, col_idx)).strip()
                for col_idx in range(ws.ncols)
            ]

            # Skip completely empty rows
            if not any(c for c in cells):
                continue

            if row_idx == 0:
                headers = cells
                continue

            row_count += 1
            formatted_row = self._format_data_row(
                headers, cells, row_count
            )
            if formatted_row:
                rows_data.append(formatted_row)

        if not rows_data:
            return None

        header_line = f"[Sheet: {ws.name}]"
        if headers and any(h for h in headers):
            header_line += f"\nColumns: {' | '.join(h for h in headers if h)}"

        return {
            "text": f"{header_line}\n" + "\n".join(rows_data),
            "row_count": row_count
        }

    def _format_data_row(
        self,
        headers: List[str],
        cells: List[str],
        row_num: int
    ) -> Optional[str]:
        """
        Format a data row as 'Header: Value' pairs.
        This creates semantically meaningful text that embeds well.

        Example:
            Row 3: Name: John Smith | Department: Engineering | Role: Senior Developer
        """
        pairs = []
        for col_idx, cell_text in enumerate(cells):
            if not cell_text or cell_text == "None":
                continue

            if col_idx < len(headers) and headers[col_idx]:
                pairs.append(f"{headers[col_idx]}: {cell_text}")
            else:
                pairs.append(cell_text)

        if not pairs:
            return None

        return f"Row {row_num}: {' | '.join(pairs)}"
